"""Document-table connector for normalized training records."""

from __future__ import annotations

import csv
import hashlib
import json
import re
from datetime import date
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from docx import Document as DocxDocument  # type: ignore[import-untyped]
from openpyxl import load_workbook  # type: ignore[import-untyped]

from complyos.connectors.base import LMSConnector
from complyos.connectors.normalization import (
    COURSE_ALIASES,
    ENROLLMENT_ALIASES,
    STATUS_SYNONYMS,
    USER_ALIASES,
    parse_date,
    parse_datetime,
    parse_float,
    remap_row,
    to_learning_status,
)
from complyos.models.domain import (
    Course,
    EmploymentStatus,
    Enrollment,
    EnrollmentStatus,
    LearningRecord,
    User,
)

DOCUMENT_IMPORT_FIELDS = (
    "id",
    "user_id",
    "course_id",
    "status",
    "assigned_date",
    "due_date",
    "completed_date",
    "completion_percentage",
    "score",
    "expires_at",
    "source_system",
    "source_record_id",
)
TRUTHY = {"true", "yes", "y", "1", "required", "mandatory"}


class DocumentExtractionError(ValueError):
    """Raised when a document cannot produce a primary table."""


def _hash_row(row: dict[str, Any]) -> str:
    payload = json.dumps(row, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _slug(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "-", value.strip().lower()).strip("-")
    return normalized or "document-item"


def _split_name(mapped: dict[str, str], fallback: str) -> tuple[str, str]:
    first = mapped.get("first_name", "")
    last = mapped.get("last_name", "")
    if first or last:
        return first, last
    full_name = mapped.get("full_name", fallback).strip()
    if not full_name:
        return "", ""
    first_name, _, last_name = full_name.partition(" ")
    return first_name, last_name


class DocumentExtractor(LMSConnector):
    """Read a Word, Excel, or CSV table into training-record models."""

    name = "document_upload"

    def __init__(
        self,
        source: str | Path | bytes | None = None,
        *,
        filename: str | None = None,
        rows: list[dict[str, Any]] | None = None,
    ) -> None:
        self.source = source
        self.filename = filename or (Path(source).name if isinstance(source, str | Path) else None)
        self._raw_rows = self._coerce_rows(rows) if rows is not None else None
        self._users: list[User] | None = None
        self._courses: list[Course] | None = None
        self._records: list[LearningRecord] | None = None
        self.skipped_rows: dict[str, int] = {"users": 0, "courses": 0, "records": 0}

    @classmethod
    def from_rows(cls, rows: list[dict[str, Any]]) -> DocumentExtractor:
        return cls(filename="rows.csv", rows=rows)

    async def authenticate(self) -> bool:
        try:
            return bool(self._load_rows())
        except DocumentExtractionError:
            return False

    async def get_users(self, filters: dict[str, Any] | None = None) -> list[User]:
        result = self._load_users()
        if filters:
            if "department" in filters:
                result = [user for user in result if user.department == filters["department"]]
            if "region" in filters:
                result = [user for user in result if user.region == filters["region"]]
            if "employment_status" in filters:
                result = [
                    user
                    for user in result
                    if user.employment_status.value == filters["employment_status"]
                ]
        return result

    async def get_courses(self, filters: dict[str, Any] | None = None) -> list[Course]:
        result = self._load_courses()
        if filters and "mandatory" in filters:
            result = [course for course in result if course.mandatory == filters["mandatory"]]
        return result

    async def get_enrollments(
        self,
        user_ids: list[str] | None = None,
        course_ids: list[str] | None = None,
    ) -> list[Enrollment]:
        result = [record.to_enrollment() for record in self._load_learning_records()]
        if user_ids:
            result = [enrollment for enrollment in result if enrollment.user_id in user_ids]
        if course_ids:
            result = [enrollment for enrollment in result if enrollment.course_id in course_ids]
        return result

    async def get_learning_records(
        self,
        user_ids: list[str] | None = None,
        course_ids: list[str] | None = None,
    ) -> list[LearningRecord]:
        result = self._load_learning_records()
        if user_ids:
            result = [record for record in result if record.user_id in user_ids]
        if course_ids:
            result = [record for record in result if record.course_id in course_ids]
        return result

    async def trigger_reminder(self, user_id: str, course_id: str) -> bool:
        return False

    def to_import_csv_text(self) -> str:
        rows = self._canonical_import_rows()
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=DOCUMENT_IMPORT_FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
        return output.getvalue()

    def _load_users(self) -> list[User]:
        if self._users is None:
            users: dict[str, User] = {}
            for row in self._load_rows():
                user_map = remap_row(row, USER_ALIASES)
                record_map = remap_row(row, ENROLLMENT_ALIASES)
                user_id = user_map.get("id") or record_map.get("user_id") or user_map.get("email")
                if not user_id:
                    self.skipped_rows["users"] += 1
                    continue
                first_name, last_name = _split_name(user_map, user_id)
                raw_status = user_map.get("employment_status", "active").lower()
                try:
                    employment_status = EmploymentStatus(raw_status)
                except ValueError:
                    employment_status = EmploymentStatus.ACTIVE
                users.setdefault(
                    user_id,
                    User(
                        id=user_id,
                        employee_id=user_map.get("employee_id", user_id),
                        email=user_map.get("email", f"{_slug(user_id)}@document.invalid"),
                        first_name=first_name,
                        last_name=last_name,
                        department=user_map.get("department", "Unknown"),
                        region=user_map.get("region", "Unknown"),
                        hire_date=parse_date(user_map.get("hire_date")) or date(1970, 1, 1),
                        employment_status=employment_status,
                        manager_id=user_map.get("manager_id"),
                        job_title=user_map.get("job_title"),
                    ),
                )
            self._users = list(users.values())
        return self._users

    def _load_courses(self) -> list[Course]:
        if self._courses is None:
            courses: dict[str, Course] = {}
            for row in self._load_rows():
                course_map = remap_row(row, COURSE_ALIASES)
                record_map = remap_row(row, ENROLLMENT_ALIASES)
                title = course_map.get("title") or record_map.get("course_id", "")
                course_id = course_map.get("id") or record_map.get("course_id") or _slug(title)
                if not course_id:
                    self.skipped_rows["courses"] += 1
                    continue
                duration = parse_float(course_map.get("duration_minutes"))
                courses.setdefault(
                    course_id,
                    Course(
                        id=course_id,
                        code=course_map.get("code", course_id),
                        title=title or course_map.get("code", course_id),
                        description=course_map.get("description"),
                        duration_minutes=int(duration) if duration else None,
                        mandatory=course_map.get("mandatory", "").lower() in TRUTHY,
                        category=course_map.get("category"),
                    ),
                )
            self._courses = list(courses.values())
        return self._courses

    def _load_learning_records(self) -> list[LearningRecord]:
        if self._records is None:
            records: list[LearningRecord] = []
            for index, row in enumerate(self._load_rows()):
                mapped = self._record_map(row)
                if not mapped.get("user_id") or not mapped.get("course_id"):
                    self.skipped_rows["records"] += 1
                    continue
                raw_status = mapped.get("status", "not_started").lower().replace(" ", "_")
                enrollment_status = STATUS_SYNONYMS.get(raw_status, EnrollmentStatus.NOT_STARTED)
                explicit_exempt = mapped.get("exempt", "").lower() in TRUTHY
                expires_at = parse_date(mapped.get("expires_at"))
                source_payload = dict(row)
                records.append(
                    LearningRecord(
                        id=mapped.get("id") or f"document-upload-{index}",
                        user_id=mapped["user_id"],
                        course_id=mapped["course_id"],
                        source_system=mapped.get("source_system") or self.name,
                        source_record_id=mapped.get("source_record_id"),
                        status=to_learning_status(
                            enrollment_status,
                            expires_at,
                            exempt=explicit_exempt,
                        ),
                        assigned_date=parse_datetime(mapped.get("assigned_date")),
                        due_date=parse_date(mapped.get("due_date")),
                        completed_date=parse_datetime(mapped.get("completed_date")),
                        completion_percentage=parse_float(mapped.get("completion_percentage"))
                        or 0.0,
                        score=parse_float(mapped.get("score")),
                        exempt=explicit_exempt or enrollment_status == EnrollmentStatus.EXEMPT,
                        expires_at=expires_at,
                        raw_source_hash=_hash_row(source_payload),
                        source_payload=source_payload,
                    )
                )
            self._records = records
        return self._records

    def _canonical_import_rows(self) -> list[dict[str, str]]:
        rows: list[dict[str, str]] = []
        for index, row in enumerate(self._load_rows()):
            mapped = self._record_map(row)
            rows.append(
                {
                    "id": mapped.get("id") or f"document-upload-{index}",
                    "user_id": mapped.get("user_id", ""),
                    "course_id": mapped.get("course_id", ""),
                    "status": mapped.get("status", "not_started"),
                    "assigned_date": mapped.get("assigned_date", ""),
                    "due_date": mapped.get("due_date", ""),
                    "completed_date": mapped.get("completed_date", ""),
                    "completion_percentage": mapped.get("completion_percentage", ""),
                    "score": mapped.get("score", ""),
                    "expires_at": mapped.get("expires_at", ""),
                    "source_system": mapped.get("source_system") or self.name,
                    "source_record_id": mapped.get("source_record_id", ""),
                }
            )
        return rows

    def _record_map(self, row: dict[str, str]) -> dict[str, str]:
        mapped = remap_row(row, ENROLLMENT_ALIASES)
        user_map = remap_row(row, USER_ALIASES)
        course_map = remap_row(row, COURSE_ALIASES)
        if "user_id" not in mapped:
            user_id = user_map.get("id") or user_map.get("email")
            if user_id:
                mapped["user_id"] = user_id
        if "course_id" not in mapped:
            title = course_map.get("title", "")
            course_id = course_map.get("id") or (title and _slug(title))
            if course_id:
                mapped["course_id"] = course_id
        return mapped

    def _load_rows(self) -> list[dict[str, str]]:
        if self._raw_rows is not None:
            return self._raw_rows
        suffix = self._suffix()
        if suffix == ".csv":
            self._raw_rows = self._read_csv_rows()
        elif suffix == ".xlsx":
            self._raw_rows = self._read_xlsx_rows()
        elif suffix == ".docx":
            self._raw_rows = self._read_docx_rows()
        else:
            raise DocumentExtractionError(
                "unsupported file type; supported types are .docx, .xlsx, and .csv"
            )
        return self._raw_rows

    def _suffix(self) -> str:
        if self.filename:
            return Path(self.filename).suffix.lower()
        if isinstance(self.source, str | Path):
            return Path(self.source).suffix.lower()
        return ""

    def _source_bytes(self) -> bytes:
        if isinstance(self.source, bytes):
            return self.source
        if isinstance(self.source, str | Path):
            try:
                return Path(self.source).read_bytes()
            except OSError as exc:
                raise DocumentExtractionError(
                    f"primary table could not be read: {exc}"
                ) from exc
        raise DocumentExtractionError("primary table source is required")

    def _read_csv_rows(self) -> list[dict[str, str]]:
        try:
            text = self._source_bytes().decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise DocumentExtractionError(f"primary table could not be read: {exc}") from exc
        reader = csv.DictReader(StringIO(text))
        if not reader.fieldnames:
            raise DocumentExtractionError("primary table is missing a header row")
        rows = self._coerce_rows(list(reader))
        if not rows:
            raise DocumentExtractionError("primary table has no data rows")
        return rows

    def _read_xlsx_rows(self) -> list[dict[str, str]]:
        try:
            workbook = load_workbook(BytesIO(self._source_bytes()), read_only=True, data_only=True)
        except Exception as exc:
            raise DocumentExtractionError(f"primary table could not be read: {exc}") from exc
        for sheet in workbook.worksheets:
            rows = [
                [_as_text(cell) for cell in row]
                for row in sheet.iter_rows(values_only=True)
                if any(_as_text(cell) for cell in row)
            ]
            if rows:
                return self._grid_to_rows(rows)
        raise DocumentExtractionError("primary table not found")

    def _read_docx_rows(self) -> list[dict[str, str]]:
        try:
            document = DocxDocument(BytesIO(self._source_bytes()))
        except Exception as exc:
            raise DocumentExtractionError(f"primary table could not be read: {exc}") from exc
        for table in document.tables:
            rows = [
                [_as_text(cell.text) for cell in row.cells]
                for row in table.rows
                if any(_as_text(cell.text) for cell in row.cells)
            ]
            if rows:
                return self._grid_to_rows(rows)
        raise DocumentExtractionError("primary table not found")

    @staticmethod
    def _grid_to_rows(grid: list[list[str]]) -> list[dict[str, str]]:
        headers = [_as_text(header) for header in grid[0]]
        if not any(headers):
            raise DocumentExtractionError("primary table is missing a header row")
        rows: list[dict[str, str]] = []
        for values in grid[1:]:
            if not any(values):
                continue
            padded = [*values, *([""] * max(0, len(headers) - len(values)))]
            rows.append(
                {
                    header: _as_text(value)
                    for header, value in zip(headers, padded, strict=False)
                    if header
                }
            )
        if not rows:
            raise DocumentExtractionError("primary table has no data rows")
        return rows

    @staticmethod
    def _coerce_rows(rows: list[dict[str, Any]]) -> list[dict[str, str]]:
        return [
            {str(key): _as_text(value) for key, value in row.items() if key is not None}
            for row in rows
        ]
